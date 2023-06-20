
import openpyxl as xl # "as xl" is to give it an aliasName to make the code shorter in using it. So openpyxl=xl

#Create wb object to load the excel spreadsheet
wb=xl.load_workbook('transactions.xlsx')

#To access the wb sheets(In this case our wb has only 1 sheet(page))
#we use sheet number to access specific sheets
sheet=wb['Sheet1'] #ensure to capitalize initials cuz this is case sensitive

#To get the cells in a sheet we use cells coordinates
cell1=sheet['a1'] #OR cell=sheet.cell(1,1)

#Get cell Value
print(cell1.value)

#To get the total rows in the sheet
print(sheet.max_row)

#to get total columns
print(sheet.max_column)

#Use a for_loop to generate the numbers 1-4 representing the rows
for row in range(1, sheet.max_row + 1): # we use +1 so as to include the 4th row
        print(row)

#To get the cell values in the 3rd column  
for row in range(2, sheet.max_row +1):  #excluding row1
        cell=sheet.cell(row,3)   #this will give use access to cell values in the rows
        print(cell.value)
#modify the values in 3rd column by multiplying by 0.9
for row in range(2, sheet.max_row +1):  
        cell=sheet.cell(row,3) 
        corrected_price=cell.value * 0.9 
        corrected_price_cell=sheet.cell(row, 4)#add a new cell to the work sheet to display corrected_price
        corrected_price_cell.value=corrected_price

#wb.save('transactions2.xlsx')  #save the changes to a different spreadsheet to monitor for bugs incase
#run the program and verify the new excel file created 

#Adding charts to the excel sheet 
print('\n# Adding charts to excel sheet')  
from openpyxl.chart import BarChart, Reference   

for row in range(2, sheet.max_row +1):  
        cell=sheet.cell(row,3) 
        corrected_price=cell.value * 0.9 
        corrected_price_cell=sheet.cell(row, 4)#add a new cell to the work sheet to display corrected_price
        corrected_price_cell.value=corrected_price
#=>Create an instance of the Reference class and store in a variable called values
values=Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4 )
#=>create an instnace of the BarChart class and store in a variable called chart
chart=BarChart()
chart.add_data(values) #calling the chart and adding data defined in values
sheet.add_chart(chart, 'e2')#add chart to sheet and define the chart coordinates on the sheet

wb.save('transactions2.xlsx') #save the changes to a different spreadsheet to monitor for bugs incase

 
#==>USING FUNCTIONS TO MAKE THE SCRIPT RE-USABLE
print('\n# using functions in managing excel sheet')
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):   #enter file name
        wb=xl.load_workbook(filename)
        sheet=wb['Sheet1']

        for row in range(2, sheet.max_row +1):  
            cell=sheet.cell(row,3) 
            corrected_price=cell.value * 0.9 
            corrected_price_cell=sheet.cell(row, 4)
            corrected_price_cell.value=corrected_price

        values=Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4 ) 
        chart=BarChart()
        chart.add_data(values) 
        sheet.add_chart(chart, 'e2') 

        wb.save(filename)  #if codes are working properly, we can override the existing file values



